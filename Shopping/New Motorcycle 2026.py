import AskAI
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


create_txt_file = True
txt_file_path = r"C:\Users\eddya\Files\Shopping\New Motorcycle 2026"
txt_file_output = []


def custom_print(*args, **kwargs):
    output = ' '.join(map(str, args))
    if create_txt_file:
        txt_file_output.append(output)
    print(output)


def get_row_with_hyperlink(file_path, index_value):
    ws = load_workbook(file_path, data_only=True)['Indexes']
    headers = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}

    index_col = headers['Index']
    brand_col = headers['Brand']
    model_col = headers['Model']
    year_col = headers['Year']
    link_col = headers['Link']

    for row in ws.iter_rows(min_row=2):
        if row[index_col - 1].value == index_value:  # -1 because Python is 0-indexed
            brand = row[brand_col - 1].value
            model = row[model_col - 1].value
            year = row[year_col - 1].value
            link_cell = row[link_col - 1]

            link = link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value

            first_string = f"{brand} {model} {year}"
            second_string = link
            return first_string, second_string
    return None, None


def get_nth_number(value, n=1):
    if not value:
        return None

    numbers = re.findall(r"\d+(?:[.,]\d+)?", value)
    if len(numbers) < n:
        return None

    num_str = numbers[n - 1].replace(",", ".")
    return float(num_str) if "." in num_str else int(num_str)


FIELD_MAP = {
    "Motorcycle": "motorcycle",
    "Seat Height [mm]": "seat height",
    "Wet Weight [kg]": "weight",
    "Engine [cc]": "engine",
    "Cylinders": "engine",
    "Power [hp]": "power",
    "Power [rpm]": "power",
    "Torque [Nm]": "torque",
    "Torque [rpm]": "torque",
    "Fuel Capacity [l]": "fuel capacity",
    "Mileage [km/l]": "mileage",
    "Price [ILS]": "price",
}


def scrape_website(url):
    response = requests.get(url)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")

    web_title = soup.select_one("h1.product_title")
    web_title = web_title.text.strip() if web_title else None
    database = {"motorcycle": web_title}

    table = soup.find("table")
    if not table:
        return None

    for row in table.find_all("tr"):
        cells = row.find_all(["th", "td"])
        if len(cells) < 2:
            continue

        key_text = cells[0].get_text(strip=True).lower()
        raw_value = cells[1].get_text(strip=True)

        if "ratio" in key_text:
            continue

        database[key_text.lower()] = raw_value
    return database


def fill_in_database(database):
    data = {}
    for field_name in FIELD_MAP.keys():
        data[field_name] = None

    for field_name, match_key in FIELD_MAP.items():
        if match_key in database and database[match_key] is not None:
            raw_value = database[match_key]
            if "Motorcycle" in field_name:
                data[field_name] = raw_value.strip()
            elif "Cylinders" in field_name:
                data[field_name] = raw_value.split("cc")[-1].strip()
            elif "rpm" in field_name.lower():
                data[field_name] = get_nth_number(raw_value, 2)
            else:
                data[field_name] = get_nth_number(raw_value)

    return data


if __name__ == "__main__":
    # ask_gemini_for_missing_data = True
    ask_gemini_for_missing_data = False
    get_motorcycle_from_excel = True
    get_motorcycle_index_number = 1

    if get_motorcycle_from_excel:
        motorcycle_excel_path = r"C:\Users\eddya\Files\Shopping\New Motorcycle 2026\New Motorcycle 2026.xlsx"
        motorcycle, url = get_row_with_hyperlink(motorcycle_excel_path, get_motorcycle_index_number)
    else:
        motorcycle = ""
        url = ""
    custom_print(f'{motorcycle = }')
    custom_print(f'{url = }')
    custom_print(f'')

    bike_database = scrape_website(url)
    bike_data = fill_in_database(bike_database)
    missing_output = ", ".join(field for field, value in bike_data.items() if value is None)
    custom_print(f'{missing_output = }')

    if ask_gemini_for_missing_data:
        response_text = AskAI.ask_gemini_motorcycle_data(motorcycle, missing_output)
        custom_print(f'{response_text = }')
        headers, values = response_text.splitlines()
        keys = headers.split(",")
        vals = values.split(",")
        bike_data.update(dict([(key, float(value) if "." in value else int(value)) for key, value in zip(keys, vals)]))

    custom_print()
    for k, v in bike_data.items():
        custom_print(f"{k}: {v}")
    custom_print()
    custom_print("\t".join(bike_data.keys()))
    custom_print("\t".join(str(v) for v in bike_data.values()) + "\n")

    if create_txt_file:
        filename = f"{txt_file_path}\\{motorcycle}.txt"
        with open(filename, 'w', encoding='utf-8') as log_file:
            log_file.write("\n".join(txt_file_output))

        print(f"\nFile '{filename}' created successfully!")
