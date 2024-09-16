import numpy as np
import pandas as pd
import os
import subprocess as sps
import sys
import ltspice
import time
import math
import datetime
import openpyxl as px
import tkinter as tk
from tkinter import filedialog
import plotly.figure_factory as ff
import plotly.io as pio
import plotly as py
from openpyxl import Workbook, load_workbook, formatting
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import threading
import shutil
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) # Ignores DeprecationWarning of LTspice addon


##################### Statistical Signals Compression Evaluation (extract from RAW file), Created by: Ido Debi (03/2021) #####################

# Class which creates thread and opens run simulation function
class myThread (threading.Thread):
   def __init__(self, threadID, filename, counter):
      threading.Thread.__init__(self)
      self.threadID = threadID
      self.filename = filename
      self.counter = counter
   def run(self):
      allowed_threads.acquire() # See if there is avalible semaphore and start the function insdie the thread
      run_simulation(self.filename, self.counter) # Run the main function
      allowed_threads.release() # Release semaphore

# Main function which run the ltspice via command line and analyze the data
def run_simulation(filename, counter): 
    global vpp_list_str
    global rms_list_str
    global vpp_list_mid
    global rms_list_mid
    global vpp_list_last
    global rms_list_last
    global vpp_list_first
    global rms_list_first

    print("\nSimulation number {} started...\n".format(counter))

    asc_file = filename+str(counter)+'.asc'

    try:
        shutil.copy(filename+'.asc', asc_file) # Copy the simulation file and rename it - simulation_file + number_of_run 
    except:
        print("\nFile {} already exists. Moving on...\n".format(iter))

    try:
        sps.run(["C:/Program Files/LTC/LTspiceXVII/XVIIx64.exe", "-b", "-run", asc_file], shell=True, check=True, stdout=sys.stdout, stderr=sps.STDOUT) # Run simulation via command line in the same process
    except:
        print("\nThere was an ERROR running simulation {}. Exiting...\n".format(counter))
        exit(-1)

    l = ltspice.Ltspice(filename+str(counter)+'.raw') # Analyze the .RAW data

    print("\nSimulation number {} ended. Analyzing Data...".format(counter))
    l.parse() # Analyze the data
    print("\nDone analyzing RAW data of simulation {}\n".format(counter))

    print("\nCalculating Vpp/Vrms from RAW data of simulation {}...\n".format(counter))
    for i in range(l.case_count): # Calculate Vrms/Vpp from the raw data
        Vstr = l.getData('V(v_str)', i)
        Vopt_mid = l.getData('V(v_opt_mid)', i)
        Vopt_first = l.getData('V(v_opt_first)', i)
        Vopt_last = l.getData('V(v_opt_last)', i)
        vpp = max(Vstr)-min(Vstr)
        vpp1 = max(Vopt_mid)-min(Vopt_mid)
        vpp2 = max(Vopt_first)-min(Vopt_first)
        vpp3 = max(Vopt_last)-min(Vopt_last)
        rms = np.sqrt(np.mean(Vstr**2))
        rms1 = np.sqrt(np.mean(Vopt_mid**2))
        rms2 = np.sqrt(np.mean(Vopt_first**2))
        rms3 = np.sqrt(np.mean(Vopt_last**2))

        threadLock.acquire() # get mutex to avoid appending from 2 threads in parallel
        vpp_list_str.append(vpp)
        rms_list_str.append(rms)
        vpp_list_mid.append(vpp1)
        rms_list_mid.append(rms1)
        vpp_list_first.append(vpp2)
        rms_list_first.append(rms2)
        vpp_list_last.append(vpp3)
        rms_list_last.append(rms3)
        threadLock.release() # release mutex

    # Remove simulation files
    os.remove(filename+str(counter)+'.asc')
    os.remove(filename+str(counter)+'.raw')
    os.remove(filename+str(counter)+'.net')
    os.remove(filename+str(counter)+'.log')
    os.remove(filename+str(counter)+'.op.raw')

    print("End of iteration {}\n".format(counter))

vpp_list_str = []
rms_list_str = []
vpp_list_mid = []
rms_list_mid = []
vpp_list_last = []
rms_list_last = []
vpp_list_first = []
rms_list_first = []

threadLock = threading.Lock() # define the mutex
threads = []
max_threads = 5 # MAX allowed threads

savename = '10MilionRuns_38HFOpts' # New file name

N = 5e6 # Number of runs. Need to be round and >= 1e5 (example: 1e6, 1e7..)

file = "C:/Users/ido.d/Desktop/Statistical_signals_compression_evaluation/LV450_Statistical_signals_compression_evaluation_38Opts" # LTSpice file to run (without .asc)

allowed_threads = threading.BoundedSemaphore(value=max_threads) # Create Semaphore to limit the number of parallel simulations

start_time = time.time() # Start of simulation time counting

print("Runing LTSpice Statistical Signals Compression Evaluation:\n")

for iter in range(1,int(N/1e5)+1): # Creates all threads to run the simulation

    threadX = myThread(iter, file, iter) # Create new Thread
    threadX.start() # Start new Thread
    threads.append(threadX) # Add threads to thread list

    time.sleep(iter*10) # wait until openning another thread

for t in threads: # Wait for all threads to complete
    t.join()

print("\nPlotting Histogram...\n") # Plotly Histogram

hist_data = [vpp_list_str, rms_list_str, vpp_list_mid, rms_list_mid,vpp_list_first, rms_list_first, vpp_list_last, rms_list_last]
group_labels = ['Vstring [Vpp]','Vstring [Vrms]', 'Vopt_middle [Vpp]','Vopt_middle [Vrms]', 'Vopt_first [Vpp]','Vopt_first [Vrms]', 'Vopt_last [Vpp]','Vopt_last [Vrms]']

fig = ff.create_distplot(hist_data, group_labels, bin_size=[1 ,0.4 ,0.025,0.01,0.025,0.01,0.025,0.01],histnorm = 'probability', show_rug=False)
fig.update_layout(title_text='LV450 - Statistical Signals Compression Evaluation [38 H1300 Optimizers is series - Amp: 3Vpp +- 0.1Vpp, Freq: 200kHz +- 10Hz, Mag: random(-180 -> 180)]', xaxis_title="Vpp / Vrms",yaxis_title="Number Of Appearances (Normalized)")
fig.update_layout()
fig.write_html(savename+'.html', auto_open=False)

# Calulate probabilities:
print("\nCreating {} file...\n".format(savename+"_probabilities.xlsx"))
wb = Workbook() # Create new xls file
ws = wb.create_sheet("Probabilities", 0)
ws['A1'] = 'X'
ws['B1'] = 'P(Vstr>X[Vpp])'
ws['C1'] = 'P(Vstr>X[Vrms])'
ws['D1'] = 'P(Vopt_mid>X[Vpp])'
ws['E1'] = 'P(Vopt_mid>X[Vrms])'
ws['F1'] = 'P(Vopt_first>X[Vpp])'
ws['G1'] = 'P(Vopt_first>X[Vrms])'
ws['H1'] = 'P(Vopt_last>X[Vpp])'
ws['I1'] = 'P(Vopt_last>X[Vrms])'
ws.column_dimensions["A"].width = 7
ws['A1'].font  = Font(b=True)
ws.column_dimensions["B"].width = 15
ws['B1'].font  = Font(b=True)
ws.column_dimensions["C"].width = 15
ws['C1'].font  = Font(b=True)
ws.column_dimensions["D"].width = 20
ws['D1'].font  = Font(b=True)
ws.column_dimensions["E"].width = 22
ws['E1'].font  = Font(b=True)
ws.column_dimensions["F"].width = 20
ws['F1'].font  = Font(b=True)
ws.column_dimensions["G"].width = 22
ws['G1'].font  = Font(b=True)
ws.column_dimensions["H"].width = 20
ws['H1'].font  = Font(b=True)
ws.column_dimensions["I"].width = 20
ws['I1'].font  = Font(b=True)

for k in range(11): # Calculate from 0.1 to 1.0 with delta of 0.1
    ws['A'+str(k+2)] = k/10
    ws['B'+str(k+2)] = len([1 for j in vpp_list_str if j >= k/10])/len(vpp_list_str)*100
    ws['C'+str(k+2)] = len([1 for j in rms_list_str if j >= k/10])/len(rms_list_str)*100
    ws['D'+str(k+2)] = len([1 for j in vpp_list_mid if j >= k/10])/len(vpp_list_mid)*100
    ws['E'+str(k+2)] = len([1 for j in rms_list_mid if j >= k/10])/len(rms_list_mid)*100
    ws['F'+str(k+2)] = len([1 for j in vpp_list_first if j >= k/10])/len(vpp_list_first)*100
    ws['G'+str(k+2)] = len([1 for j in rms_list_first if j >= k/10])/len(rms_list_first)*100
    ws['H'+str(k+2)] = len([1 for j in vpp_list_last if j >= k/10])/len(vpp_list_last)*100
    ws['I'+str(k+2)] = len([1 for j in rms_list_last if j >= k/10])/len(rms_list_last)*100

    if k == 10: # Calculate P(Vstr) from 2 to 30 with delta of 1
        for t in range(2,31,1):
            ws['A'+str(t+11)] = t
            ws['B'+str(t+11)] = len([1 for j in vpp_list_str if j >= t])/len(vpp_list_str)*100
            ws['C'+str(t+11)] = len([1 for j in rms_list_str if j >= t])/len(rms_list_str)*100

for rows in ws.iter_rows(): # Moving throught every cell at the active sheet in order to apply styling
    for cell in rows:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

###################################################################################################

ws1 = wb.create_sheet("Rayleigh distribution", 1)
ws1['B1'] = 'Maximum Likelihood Estimation'
ws1['C1'] = 'Mean'
ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 15
ws1['B1'].font  = Font(b=True)
ws1.column_dimensions["C"].width = 15
ws1['C1'].font  = Font(b=True)


ws1['A2'] = 'Vstr>X[Vpp]'
ws1['A3'] = 'Vstr>X[Vrms]'
ws1['A4'] = 'Vopt_mid>X[Vpp]'
ws1['A5'] = 'Vopt_mid>X[Vrms]'
ws1['A6'] = 'Vopt_first>X[Vpp]'
ws1['A7'] = 'Vopt_first>X[Vrms]'
ws1['A8'] = 'Vopt_last>X[Vpp]'
ws1['A9'] = 'Vopt_last>X[Vrms]'

def turn_to_power(list, power=2): # Function which returns the same list with numbers^2
    return [number**power for number in list]

# calculate Maximum Likelihood Estimation + Mean
MLE1 = np.sqrt((np.sum(turn_to_power(vpp_list_str)))/(2*N))
MEAN1 = MLE1*np.sqrt(math.pi/2)
ws1['B2'] = MLE1
ws1['C2'] = MEAN1
MLE2 = np.sqrt((np.sum(turn_to_power(rms_list_str)))/(2*N))
MEAN2 = MLE2*np.sqrt(math.pi/2)
ws1['B3'] = MLE2
ws1['C3'] = MEAN2
MLE3 = np.sqrt((np.sum(turn_to_power(vpp_list_mid)))/(2*N)) 
MEAN3 = MLE3*np.sqrt(math.pi/2)
ws1['B4'] = MLE3
ws1['C4'] = MEAN3
MLE4 = np.sqrt((np.sum(turn_to_power(rms_list_mid)))/(2*N))
MEAN4 = MLE4*np.sqrt(math.pi/2)
ws1['B5'] = MLE4
ws1['C5'] = MEAN4
MLE5 = np.sqrt((np.sum(turn_to_power(vpp_list_first)))/(2*N))
MEAN5 = MLE5*np.sqrt(math.pi/2)
ws1['B6'] = MLE5
ws1['C6'] = MEAN5
MLE6 = np.sqrt((np.sum(turn_to_power(rms_list_first)))/(2*N))
MEAN6 = MLE6*np.sqrt(math.pi/2)
ws1['B7'] = MLE6
ws1['C7'] = MEAN6
MLE7 = np.sqrt((np.sum(turn_to_power(vpp_list_last)))/(2*N))
MEAN7 = MLE7*np.sqrt(math.pi/2)
ws1['B8'] = MLE7
ws1['C8'] = MEAN7
MLE8 = np.sqrt((np.sum(turn_to_power(rms_list_last)))/(2*N))
MEAN8 = MLE8*np.sqrt(math.pi/2)
ws1['B9'] = MLE8
ws1['C9'] = MEAN8

for rows in ws1.iter_rows(): # Moving throught every cell at the active sheet in order to apply styling
    for cell in rows:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

##########################################################################################################

wb.save(savename+"_probabilities.xlsx") # Save xls file
print("Done!\n")

os.startfile(savename+"_probabilities.xlsx") # open the xlsx file 

elapsed_time = time.time() - start_time # End of simulation time counting
print("\n--> The process took: %s (Hours,Min,Sec)\n" % (datetime.timedelta(seconds=math.floor(elapsed_time)))) # Print the simulation elapsed time