import numpy
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import numpy as np
import SPI_Reading
import math
from scipy import signal
import pandas as pd
import openpyxl
from openpyxl import load_workbook




def iir_lpf(data_wind,cutoff, fs, order):
    B, A = butter_lowpass1(cutoff, fs, order=order)
    b0 = B[0]
    b1 = B[1]
    a0 = A[0] # Almost always 1
    a1 = A[1]#(b0+b1) -1
    x_n_1 = data_wind[0]
    Y_IIR_1 = x_n_1 #0#750
    yVec = [Y_IIR_1]
    for index, value in enumerate(data_wind[1:]):
        x_n = data_wind[index]
        Y_IIR = (b0 * x_n +b1 * x_n_1 - (a1) * Y_IIR_1)/a0
        yVec.append(Y_IIR)
        Y_IIR_1 = Y_IIR
        x_n_1 = x_n
    return yVec

def Vdciir(Vdc_fast_With_arc,Fs_SPI,time_spi_fast,Name,orderFilter ,fc ):
    Window_size = int(math.ceil(Fs_SPI/50))
    timeVDC_SlowVec = []
    FilteredSignalLPInitalize_Vec = []
    for index in range(0, int(Window_size*2) ,int(Window_size)):
        ind = index
        Vdc_Window = Vdc_fast_With_arc[ind:ind + Window_size]
        FilteredSignalLPInitalize = iir_lpf(Vdc_Window,fc, Fs_SPI, order=orderFilter) # iir_lpf(data_wind,cutoff, fs, order)
        FilteredSignalLPInitalize_Vec.extend(FilteredSignalLPInitalize)

    FilteredSignalLPInitalize_VecCUT = FilteredSignalLPInitalize_Vec[int(Window_size / 2):int(Window_size*1.2)]
    FirstValMin = min(FilteredSignalLPInitalize_VecCUT)
    FirstIndxMin = FilteredSignalLPInitalize_VecCUT.index(FirstValMin) + int(Window_size / 2)
    ValVdc_WideWindowMinVec2 = []
    IndxVdc_WideWindowMinVec = []
    FilteredSignalLP_Vec = []
    time_FilteredSignalLP_Vec = time_spi_fast[FirstIndxMin:]
    ExccedTHCounterVecF =[]
    LowerThVec = []
    k = 0
    Cyc_n = 0
    diffB = 0
    diffBVec = []
    LowerThIndxVec = []
    ExccedTHCounterVecFIndx = []
    diffBVecIndx = []
    VdcMeanVec = []
    VdcMeanVecIndx = []
    VdcAvgdiffVec = []
    VdcAvgdiffVecIndx = []
    MultiplyVec = []
    ValVdc_WideWindowMinB = 0
    ValVdc_WideWindowMinE = 0
    LowerVec = []
    LowerVecIndx = []
    # Extract locals minimums per cycle
    # for index in range( int(FirstIndxMin), len(Vdc_fast_With_arc),int(Window_size)):
    for Index in range( int(FirstIndxMin), len(Vdc_fast_With_arc),int(Window_size)):
        timeSlow = time_spi_fast[Index]
        timeVDC_SlowVec.append(timeSlow)
        ind = Index
        k = k +1
        Cyc_n = Cyc_n + 1

        Vdc_WideWindow = Vdc_fast_With_arc[ind:ind + Window_size]

        try:
            Vdc_WideWindowFiltered = iir_lpf(Vdc_WideWindow, fc, Fs_SPI, order=orderFilter)

        except:
            break

        Vdc_WideWindowFilteredList = Vdc_WideWindowFiltered#.tolist()
        FilteredSignalLP_Vec.extend(Vdc_WideWindowFilteredList)
        if Index == int(FirstIndxMin):
            ValVdc_WideWindowMinB = Vdc_WideWindowFilteredList[0]
            ValVdc_WideWindowMinE = min(Vdc_WideWindowFilteredList[int(Window_size*0.7):])
            ValVdc_WideWindowMinEIndx = Vdc_WideWindowFilteredList.index(ValVdc_WideWindowMinE)+ Index
            ValVdc_WideWindowMinBIndx = Vdc_WideWindowFilteredList.index(ValVdc_WideWindowMinB)+ Index

        else:
            ValVdc_WideWindowMinB = min(Vdc_WideWindowFilteredList[:int(Window_size*0.3)])#,ValVdc_WideWindowMinE])
            ValVdc_WideWindowMinBIndx = Vdc_WideWindowFilteredList.index(ValVdc_WideWindowMinB) + Index

            if ValVdc_WideWindowMinB < ValVdc_WideWindowMinE:
                ValVdc_WideWindowMinBIndx = ValVdc_WideWindowMinBIndx
                ValVdc_WideWindowMinB = ValVdc_WideWindowMinB
            else:
                ValVdc_WideWindowMinBIndx = ValVdc_WideWindowMinEIndx
                ValVdc_WideWindowMinB = ValVdc_WideWindowMinE
            try:
                ValVdc_WideWindowMinE = min(Vdc_WideWindowFilteredList[int(Window_size*0.7):])#min([Vdc_WideWindowFilteredList[:int(Window_size/2-15)],ValVdc_WideWindowMinE])
                ValVdc_WideWindowMinEIndx = Vdc_WideWindowFilteredList.index(ValVdc_WideWindowMinE)+ Index
            except:
                ValVdc_WideWindowMinE = Vdc_WideWindowFilteredList[-1]#min(Vdc_WideWindowFilteredList[int(Window_size/2):])#min([Vdc_WideWindowFilteredList[:int(Window_size/2-15)],ValVdc_WideWindowMinE])
                ValVdc_WideWindowMinEIndx = Vdc_WideWindowFilteredList.index(ValVdc_WideWindowMinE)+ Index
        ValVdc_WideWindowMinVec2.append(ValVdc_WideWindowMinB)
        IndxVdc_WideWindowMinVec.append(ValVdc_WideWindowMinBIndx)


        ExccedTHCounterF = 0

        VdcAvg = np.mean(Vdc_WideWindow)
        VdcMeanVec.extend([VdcAvg])
        VdcMeanVecIndx.extend([Index+Window_size])

        if len(ValVdc_WideWindowMinVec2) >= 5:

            # VdcAvgdiff1 = (VdcMeanVec[-2] - VdcMeanVec[-3]) #*(2/3) #/2#####################
            # VdcAvgdiff2 = (VdcMeanVec[-2] - VdcMeanVec[-4])  *(2/3) #/2
            # VdcAvgdiff3 = (VdcMeanVec[-2] - VdcMeanVec[-5])  *(3/4)#*(2/3) #/2
            VdcAvgdiff1 = (VdcMeanVec[-2] - VdcMeanVec[-3]) *3/4#*(2/3) #/2#####################
            VdcAvgdiff2 = (VdcMeanVec[-2] - VdcMeanVec[-4]) *3/4# *(2/3) #/2
            VdcAvgdiff3 = (VdcMeanVec[-2] - VdcMeanVec[-5]) *3/4# *(2/3)#*(2/3) #/2

            VdcAvgdiff11 = float(0 if VdcAvgdiff1 > 0 else -VdcAvgdiff1)
            VdcAvgdiff22 = float(0 if VdcAvgdiff2 > 0 else -VdcAvgdiff2)
            VdcAvgdiff33 = float(0 if VdcAvgdiff3 > 0 else -VdcAvgdiff3)
            VdcAvgdiff13 = float(1 if VdcAvgdiff11 + VdcAvgdiff33 > 1 else VdcAvgdiff11 + VdcAvgdiff33)
            VdcAvgdiff32 = float(1 if VdcAvgdiff33 + VdcAvgdiff22 > 1 else VdcAvgdiff33 + VdcAvgdiff22)
            VdcAvgdiff12 = float(1 if VdcAvgdiff11 + VdcAvgdiff22 > 1 else VdcAvgdiff11 + VdcAvgdiff22)
            VdcAvgdiff13_32_max = np.max([VdcAvgdiff13,VdcAvgdiff32,VdcAvgdiff12])
            VdcAvgdiff = VdcAvgdiff13_32_max**2#VdcAvgdiff12
            # VdcAvgdiff13_32_max = np.max([VdcAvgdiff13,VdcAvgdiff32])
            # VdcAvgdiff = VdcAvgdiff13_32_max*VdcAvgdiff12

            VdcAvgdiffVecIndx.extend([Index + Window_size])

            # Lower1 = (ValVdc_WideWindowMinVec2[-1] - ValVdc_WideWindowMinVec2[-2])*(2/3) #/2/2
            Lower1 = (ValVdc_WideWindowMinVec2[-1] - ValVdc_WideWindowMinVec2[-2])*(3/4)#*(2/3)#*(2/3) #/2/2
            Lower2 =  (ValVdc_WideWindowMinVec2[-1] - ValVdc_WideWindowMinVec2[-3])*(3/4)#*(2/3)#*(2/3) #/2/2
            Lower3 =  (ValVdc_WideWindowMinVec2[-1] - ValVdc_WideWindowMinVec2[-4])#*(3/4)#*(2/3) #*(1/2) #/2/2###############
            Lower11 = float(0 if Lower1 > 0 else -Lower1)
            Lower22 = float(0 if Lower2 > 0 else -Lower2)
            Lower33 = float(0 if Lower3 > 0 else -Lower3)
            Lower13 = float(1 if Lower11 + Lower33 > 1 else Lower11 + Lower33)
            Lower32 = float(1 if Lower33 + Lower22 > 1 else Lower33 + Lower22)
            Lower12 = float(1 if Lower11 + Lower22 > 1 else Lower11 + Lower22)
            Lower13_32_max = np.max([Lower13,Lower32,Lower12])
            Lower = Lower13_32_max*Lower12
            # Lower13_32_max = np.max([Lower13,Lower32])
            # Lower = Lower13_32_max*Lower12


            LowerVec.append(Lower)
            LowerVecIndx.extend([int(Index + Window_size)])

            LowerThC = np.min(np.array(ValVdc_WideWindowMinVec2[-5:-2])) - (np.max(np.array(ValVdc_WideWindowMinVec2[-5:-2]))-np.min(np.array(ValVdc_WideWindowMinVec2[-5:-3])))  #2*np.std(np.array(ValVdc_WideWindowMinVec2[-5:-3]))
            checkCycle = FilteredSignalLP_Vec[int(-2*Window_size):int(-Window_size)]#
            LowerThVec.extend([LowerThC]*Window_size)
            LowerThIndxVec.extend(list(range(int(Index-Window_size ), int(Index ))))


            for value in checkCycle:
                if value < LowerThC:
                    ExccedTHCounterF += 1
            ExccedTHCounterVecFIndx.extend([Index + Window_size])
            ########################
            diffB1 = -((ValVdc_WideWindowMinVec2[-1]) - np.max(ValVdc_WideWindowMinVec2[-5:-2]))*3/4#*3/4 #np.mean(LowerThVec[-int(5*Window_size):])
            diffB2 = -((ValVdc_WideWindowMinVec2[-2]) - np.mean(ValVdc_WideWindowMinVec2[-5:-3]))*3/4 #np.mean(LowerThVec[-int(5*Window_size):])
            diffB11 = float(0 if (diffB1 ) < 0 else (diffB1)) # last change
            diffB22 = float(0 if (diffB2 ) < 0 else (diffB2))#*3/4 )) # was 2/3###############################
            diffB12 = float(0 if (diffB1 +diffB2) < 0.3 else (diffB1 +diffB2))
            diffB = float(1 if diffB12 > 1 else (diffB12)**2)
            ###################
            # diffB1 = -((ValVdc_WideWindowMinVec2[-1]) - np.mean(ValVdc_WideWindowMinVec2[-5:-3])) #np.mean(LowerThVec[-int(5*Window_size):])
            # diffB2 = -((ValVdc_WideWindowMinVec2[-2]) - np.mean(ValVdc_WideWindowMinVec2[-5:-3])) #np.mean(LowerThVec[-int(5*Window_size):])
            # diffB11 = float(0 if (diffB1 ) < 0 else (diffB1)) # last change
            # diffB22 = float(0 if (diffB2 ) < 0 else (diffB2))#*3/4 )) # was 2/3###############################
            # diffB12 = float(0 if (diffB11 +diffB22) < 0.25 else (diffB11 +diffB22))
            # diffB = float(1 if diffB12 > 0.95 else (diffB12)**2)

            if ExccedTHCounterF >= 1:  #
                ExccedTHCounterVecF.extend([ExccedTHCounterF/(Window_size/2)])
            else:
                ExccedTHCounterVecF.extend([0])

            if VdcAvgdiff > 0:
                VdcAvgdiffVec.extend([VdcAvgdiff])
            else:
                VdcAvgdiff = 0
                VdcAvgdiffVec.extend([0])

            if diffB > 0 :
                diffBVec.extend([diffB])
                diffBVecIndx.extend([Index+Window_size])
                ExccedTHCounter = ExccedTHCounterF / (Window_size / 2)
                Multiply = diffB * ExccedTHCounter * VdcAvgdiff
                MultiplyVec.extend([Multiply])

            else:
                diffBVec.extend([0])
                diffBVecIndx.extend([Index+Window_size])
                MultiplyVec.extend([0])


    return FilteredSignalLP_Vec ,time_FilteredSignalLP_Vec , ValVdc_WideWindowMinVec2, IndxVdc_WideWindowMinVec , ExccedTHCounterVecF , ExccedTHCounterVecFIndx,LowerThVec,LowerThIndxVec ,diffBVec,diffBVecIndx ,VdcAvgdiffVec,VdcAvgdiffVecIndx,MultiplyVec,LowerVec,LowerVecIndx

def Counter_IAC_RMS_calc(IAC_RMS,timeRmsIac,buffer_size = 50,windowSize = 20, filter_size=15,TH_Iac = 0.2):
    IacCounterVector = []
    timeCounterVec = []
    for index in range(buffer_size, len(IAC_RMS)):
        IacCounter = 0
        ind = index - buffer_size
        timeCounter = timeRmsIac[index]
        timeCounterVec.append(timeCounter)
        # IAC_RMS Fill buff
        bufferI = IAC_RMS[ind:ind + buffer_size]
        windowI = bufferI[-windowSize:]
        filterI = bufferI[-(windowSize+filter_size): -windowSize]


        AVG_in_filter_I = np.mean(filterI)

        for value in windowI:
            if value < (AVG_in_filter_I- TH_Iac):
                IacCounter+=1
        IacCounterVector +=[IacCounter]

    return IacCounterVector,timeCounterVec

def initialize_fig(row=4, col=1, plots_per_pane=4, shared_xaxes=True, subplot_titles=['Rx', 'Rx', 'Rx', 'Rx']):
    all_specs = np.array([[{"secondary_y": True}] for x in range((row * col))])
    all_specs_reshaped = (np.reshape(all_specs, (col, row)).T).tolist()
    fig = make_subplots(rows=row, cols=col, specs=all_specs_reshaped, shared_xaxes=shared_xaxes,
                        subplot_titles=subplot_titles)

    return fig

def df_time_reset(dfx, x_col):
    """
        Function resets data chunk time vector
        `dfx`   - Pandas Data Frame
        `x_col` - exact Time col name
        Example of usage :
            df1 = df_time_reset(df1,'Time')
    """
    prop_df = dfx.copy(deep=True)
    prop_df[x_col] = prop_df[x_col] - dfx[x_col].iloc[0]
    return prop_df

def Counter_Energy_calcCorrect(log_energy,timeEnergy,buffer_size = 50,windowSize = 20, filter_size=15, TH_Energy = 12):
    EnergyCounterVector = []
    timeCounterVec = []
    for index in range(buffer_size, len(log_energy)):
        EnergyCounter = 0
        ind = index - buffer_size
        timeCounter = timeEnergy[index]
        timeCounterVec.append(timeCounter)
        # Energy Fill buff
        bufferE = log_energy[ind:ind + buffer_size]
        windowE = bufferE[-windowSize:]
        filterE = bufferE[-(windowSize+filter_size): -windowSize]
        min_filterEnergy = np.min(filterE)

        for value in windowE:
            if value >= (min_filterEnergy+ TH_Energy):
                EnergyCounter+=1
        EnergyCounterVector +=[EnergyCounter]
    return EnergyCounterVector,timeCounterVec

def Counter_Iac_calcCorrect(RmsIac,time_RmsIac,buffer_size = 50,windowSize = 20, filter_size=15, TH_Iac = 0.2):
    IacCounterVector = []
    timeCounterVec = []
    for index in range(buffer_size, len(RmsIac)):
        IacCounter = 0
        ind = index - buffer_size
        timeCounter = time_RmsIac[index]
        timeCounterVec.append(timeCounter)
        # Iac Fill buff
        bufferI = RmsIac[ind:ind + buffer_size]
        windowI = bufferI[-windowSize:]
        filterI = bufferI[-(windowSize+filter_size): -windowSize]
        AVG_in_filter_I = np.mean(filterI)

        for value in windowI:
            if value < (AVG_in_filter_I- TH_Iac):
                IacCounter+=1
        IacCounterVector +=[IacCounter]
    return IacCounterVector,timeCounterVec

def energy_rise_calcCorrect(log_energy, timeEnergy, windowSize=20, filter_size=15, samTH=12, buffer_size=50):
    VecThresh = []
    timeWindVec = []
    for index in range(buffer_size, len(log_energy)-1):
        timeWind = timeEnergy[index ]
        timeWindVec.append(timeWind)
        ind = index - buffer_size
        buffer1 = log_energy[ind:ind + buffer_size]
        filter1 = buffer1[-(windowSize + filter_size): -windowSize]
        window1 = buffer1[-windowSize:]
        min_in_filter = min(filter1)
        SortWind = numpy.sort(window1)
        SortWindFlip = np.flip(SortWind)
        if len(SortWindFlip) > samTH:
            K_sampValWin = SortWindFlip[samTH - 1]
            VecThresh += [K_sampValWin - min_in_filter]
        if (len(SortWindFlip) <= samTH) & (SortWindFlip.size > 0):
            K_sampValWin = SortWindFlip[0]
            VecThresh += [K_sampValWin - min_in_filter]
        index += 1
    VecThresh = [0 if x < 0 else float(x) for x in VecThresh]
    return VecThresh, timeWindVec

def VDC_Drop_calc(VDC_Slow,timeS, windowSize = 20, filter_size=15, samTH = 12):
    buffer_size = 50 # 50
    min_in_filterEachbuffer = []
    VecThresh = []
    timeWindVec = []
    for index in range(buffer_size, len(VDC_Slow)):
        timeWind = timeS[index]
        timeWindVec.append(timeWind)
        ind = index - buffer_size
        buffer1 = VDC_Slow[ind:ind + buffer_size]
        window1 = buffer1[-windowSize:]
        filter1 = buffer1[-(windowSize+filter_size): -windowSize]
        min_in_filter = np.mean(filter1)
        min_in_filterEachbuffer += [min_in_filter]
        SortWind = numpy.sort(window1)
        SortWindFlip = SortWind
        if len(SortWindFlip) > samTH:
            K_sampValWin = SortWindFlip[samTH-1]
            VecThresh += [K_sampValWin - min_in_filter]
        if (len(SortWindFlip) <= samTH) & (SortWindFlip.size > 0):
            K_sampValWin = SortWindFlip[0]
            VecThresh += [K_sampValWin - min_in_filter]
        index += 1
    VecThresh = [0 if x > 0 else float(x) for x in VecThresh]
    return np.abs(VecThresh),timeWindVec

def stage1_energy_calc50_ZCR(SignalAfterDownsample ,TimeAfterDownsample,FsAfterDownsample, zero_crossingsIdx2,fIf = 6000,alpha = 0.2121,TH=12,windowSize=20,filter_size=15,samTH=12):
    ARC_MAX_BUFFER_SIZE = 50
    Signal = np.array(SignalAfterDownsample, dtype=float)
    DClevel = 0# 8200
    DcLevelCounter = 0
    DCLevelSum = 0
    PlaceInBit = 0
    SamplesPerBitFzcr = np.diff(zero_crossingsIdx2) #math.floor(FsAfterDownsample / arcbw)
    zcrlast = len(Signal) - zero_crossingsIdx2[-1]
    SamplesPerBitF11 = np.insert(SamplesPerBitFzcr,0,zero_crossingsIdx2[0]) #zero_crossingsIdx2[0]
    SamplesPerBitF = np.append(SamplesPerBitF11,zcrlast) #zero_crossingsIdx2[0]

    HammWindow = []
    for j in range(len(SamplesPerBitF)):
        HammWindowj = []
        for i in range(SamplesPerBitF[j]):
            val = 0.54 - 0.46 * np.cos(2 * np.pi * i / SamplesPerBitF[j])
            HammWindowj.append(val)
        HammWindow.append(HammWindowj)

    # build dft vectors

    EnergyReSum = 0
    EnergyImSum = 0
    LastEnergy = 0
    EnergyVec = []
    m = 0
    CurrentIndex = 0
    count = 0
    After_First_buffer = False
    ArcDetectStage1 = False
    # LastEnergy = 0
    Buffer = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
              0, 0,
              0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    ArcInd = []

    # timeDow = np.array(TimeAfterDownsample[zero_crossingsIdx2[0]:zero_crossingsIdx2[-1]], dtype=float)

    # start
    k = 1 # try k =1 until -1 and substract drom signal
    for index, value in enumerate(Signal[zero_crossingsIdx2[0]:zero_crossingsIdx2[-1]]):
        DCLevelSum += value
        sample = value - DClevel
        DcLevelCounter += 1
        if DcLevelCounter == SamplesPerBitF[k]: # to check
            DClevel = DCLevelSum / DcLevelCounter
            DCLevelSum = 0
            DcLevelCounter = 0
        PlaceInBit += 1
        if PlaceInBit >= SamplesPerBitF[k]:  # each 357 do calc
            PlaceInBit = 0
            k+=1
            Energy = np.square(EnergyReSum) + np.square(EnergyImSum)
            EnergyReSum = 0
            EnergyImSum = 0
            EnergyIIR = alpha * Energy + (1 - alpha) * LastEnergy
            EnergyVec.append(EnergyIIR)
            LastEnergy = EnergyIIR
            Buffer[CurrentIndex] = EnergyIIR
            CurrentIndex += 1
            if CurrentIndex == ARC_MAX_BUFFER_SIZE - 1:
                CurrentIndex = 0
                After_First_buffer = True
            if After_First_buffer:
                Filter = Buffer[:filter_size]
                window = Buffer[filter_size:filter_size + windowSize]
                min_in_filter = min(Filter)
                m += 1
                for val_in_window in window:
                    if val_in_window - TH > min_in_filter:
                        count += 1
                    if count > samTH:
                        ArcDetectStage1 = True
        else:
            sample_after_window = HammWindow[k][PlaceInBit] * sample
            EnergyImSum += (sample_after_window * np.sin(2 * np.pi * fIf * PlaceInBit / FsAfterDownsample))
            EnergyReSum += (sample_after_window * np.cos(2 * np.pi * fIf * PlaceInBit / FsAfterDownsample))


    log_energy1 = 10 * np.log10([e / 100000 for e in EnergyVec])
    timeDow = np.array(TimeAfterDownsample[zero_crossingsIdx2[1:]], dtype=float)
    timeE = timeDow
    timeEnergy = timeE[1:]
    log_energy = log_energy1[1:]

    return log_energy,timeEnergy

def stage1_energy_calc50_ZCR_hammingFixed(SignalAfterDownsample ,TimeAfterDownsample,FsAfterDownsample, zero_crossingsIdx2,fIf = 6000,alpha = 0.2121,TH=12,windowSize=20,filter_size=15,samTH=12):
    ARC_MAX_BUFFER_SIZE = 50
    Signal_zcr = np.array(SignalAfterDownsample[zero_crossingsIdx2[0]:zero_crossingsIdx2[-1]], dtype=float)
    DClevel = 0
    DcLevelCounter = 0
    DCLevelSum = 0
    PlaceInBit = 0
    SamplesPerBitFzcr = np.diff(zero_crossingsIdx2)
    SamplesPerBitF = SamplesPerBitFzcr

    #build fixed Hamming
    HammWindow = []
    for i in range(0,333):
        val = 0.54 - 0.46 * np.cos(2 * np.pi * i / 333)
        HammWindow.append(val)

    # build dft vectors

    EnergyReSum = 0
    EnergyImSum = 0
    LastEnergy = 0
    EnergyVec = []
    CurrentIndex = 0
    Buffer = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
              0, 0,
              0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    # start
    k = 0
    for index, value in enumerate(Signal_zcr):
        DCLevelSum += value
        sample = value - DClevel
        DcLevelCounter += 1
        if DcLevelCounter == SamplesPerBitF[k]:
            DClevel = DCLevelSum / DcLevelCounter
            DCLevelSum = 0
            DcLevelCounter = 0
        PlaceInBit += 1
        if PlaceInBit >= SamplesPerBitF[k]:
            PlaceInBit = 0
            k+=1
            Energy = np.square(EnergyReSum) + np.square(EnergyImSum)
            EnergyReSum = 0
            EnergyImSum = 0
            EnergyIIR = alpha * Energy + (1 - alpha) * LastEnergy
            EnergyVec.append(EnergyIIR)
            LastEnergy = EnergyIIR
            Buffer[CurrentIndex] = EnergyIIR
            CurrentIndex += 1
            if CurrentIndex == ARC_MAX_BUFFER_SIZE - 1:
                CurrentIndex = 0
        else:
            try:
                sample_after_window = HammWindow[PlaceInBit] * sample
            except:
                sample_after_window = HammWindow[-1] * sample
            EnergyImSum += (sample_after_window * np.sin(2 * np.pi * fIf * PlaceInBit / FsAfterDownsample))
            EnergyReSum += (sample_after_window * np.cos(2 * np.pi * fIf * PlaceInBit / FsAfterDownsample))


    log_energy1 = 10 * np.log10([e / 100000 for e in EnergyVec])
    timeDow = np.array(TimeAfterDownsample[zero_crossingsIdx2[:-1]], dtype=float)
    # timeDow = np.array(TimeAfterDownsample[zero_crossingsIdx2[0:]], dtype=float)
    timeE = timeDow
    timeEnergy = timeE#[1:]
    log_energy = log_energy1#[1:]

    return log_energy,timeEnergy

def zcr_calc(Iac_L1C):
    # ZCR Iac L1 - I.e., zero_crossings will contain the indices of elements before which a zero crossing occurs. If you want the elements after, just add 1 to that array.
    zero_crossings = numpy.where(numpy.diff(numpy.sign(Iac_L1C)))[0]  # .astype(int)
    zero_crossingsIdx = zero_crossings.astype(int)
    if Iac_L1C[zero_crossingsIdx[0] + 2] > 0:
        zero_crossingsIdx22 = np.array(zero_crossingsIdx[::2]).astype(int)  # [random_values.astype(int)]#zero_crossingsIdx[::2]
    else:
        zero_crossingsIdxP = zero_crossingsIdx[1:]
        zero_crossingsIdx22 = np.array(zero_crossingsIdxP[::2]).astype(int)
    DiffZCR = np.diff(zero_crossingsIdx22)>300 #zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    zero_crossingsIdx2_1 = zero_crossingsIdx22[1:]
    zero_crossingsIdx2_lst = zero_crossingsIdx2_1[DiffZCR].tolist()
    zero_crossingsIdx2_lst.insert(0,zero_crossingsIdx22[0])
    zero_crossingsIdx2 = np.array(zero_crossingsIdx2_lst)
    return zero_crossingsIdx2

def butter_lowpass1(cutoff, fs, order):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = signal.butter(order, normal_cutoff, btype='low', analog=False)
    # b, a = signal.butter(order, normal_cutoff, btype='low', analog=True)
    return b, a

def newVdcProcessCounterZCR(VdcFast,time_RX_RAWC,zero_crossingsIdx2,cutoff,fs,order,windowSize = 10, filter_size=25, samTH = 2, TH_Vdc = 0.05):
    B, A = butter_lowpass1(cutoff, fs, order=order)
    b0 = B[0]
    b1 = B[1]
    a0 = A[0] # Almost always 1
    a1 = A[1]#(b0+b1) -1
    x_n_1 = VdcFast[0]
    Y_IIR_1 = x_n_1
    VdcVec = [Y_IIR_1]
    for index in range(0, int(len(VdcFast)-1)):
        x_n = VdcFast[index]
        Y_IIR = (b0 * x_n +b1 * x_n_1 - (a1) * Y_IIR_1)/a0
        VdcVec.append(Y_IIR)
        Y_IIR_1 = Y_IIR
        x_n_1 = x_n

    T = 1 / fs
    N_ = len(VdcVec)
    time_ = np.linspace(0, N_ * T, N_)
    VdcVecMean = []
    for i in range(0, len(zero_crossingsIdx2)-1):
        bufferVdc = VdcVec[zero_crossingsIdx2[i]:zero_crossingsIdx2[i+1]]
        MeanVdc = np.mean(bufferVdc)
        VdcVecMean.append(MeanVdc)


    time = np.array(time_RX_RAWC[zero_crossingsIdx2[:-1]], dtype=float)
    # time = np.array(time_RX_RAWC[zero_crossingsIdx2[0:]], dtype=float)

    ##################################
    buffer_size = 50
    VdcCounterVector =  []
    timeCounterVecVdc = []
    for index in range(buffer_size, len(VdcVecMean)):
        VdcCounter = 0
        ind = index - buffer_size
        timeCounterVdc = time[index]
        timeCounterVecVdc.append(timeCounterVdc)
        # VDC_Slow Fill buff
        bufferV = VdcVecMean[ind:ind + buffer_size]
        windowV = bufferV[-windowSize:]
        filterV = bufferV[-(windowSize+filter_size): -windowSize]
        Min_in_filter_V = np.min(filterV)
        for value in windowV:
            if value < (Min_in_filter_V- TH_Vdc):
                VdcCounter+=1
        VdcCounterVector +=[VdcCounter]

    #####################################
    buffer_size = 50
    min_in_filterEachbuffer = []
    VecThresh = []
    timeWindVec = []
    for index in range(buffer_size, len(VdcVecMean)):
        timeWind = time[index]
        timeWindVec.append(timeWind)
        ind = index - buffer_size
        buffer1 = VdcVecMean[ind:ind + buffer_size]
        window1 = buffer1[-windowSize:]
        filter1 = buffer1[-(windowSize + filter_size): -windowSize]
        min_in_filter = np.min(filter1)
        min_in_filterEachbuffer += [min_in_filter]
        SortWind = numpy.sort(window1)
        SortWindFlip = SortWind
        if len(SortWindFlip) > samTH:
            K_sampValWin = SortWindFlip[samTH - 1]
            VecThresh += [K_sampValWin - min_in_filter]
        if (len(SortWindFlip) <= samTH) & (SortWindFlip.size > 0):
            K_sampValWin = SortWindFlip[0]
            VecThresh += [K_sampValWin - min_in_filter]
        index += 1
    VecThresh = [0 if x > 0 else float(x) for x in VecThresh]

    return VdcVec, time_,np.abs(VecThresh),timeWindVec , VdcCounterVector , timeCounterVecVdc,VdcVecMean,time

def IacProcessZCR(Iac,time_RX_RAWC,ZCRindx2):
    RmsIacVEC = []
    for i in range(0, len(ZCRindx2)-1):
        bufferIac = Iac[ZCRindx2[i]:ZCRindx2[i+1]]
        RmsIac = np.sqrt(np.mean(np.array(bufferIac) ** 2))
        RmsIacVEC.append(RmsIac)

    time = np.array(time_RX_RAWC[ZCRindx2[:-1]], dtype=float)
    return RmsIacVEC, time

def E_V_I_FlagDetection(EnergyCounterVectorBinary,IacCounterVectorBinary,VdcCounterVectorBinary):
    if np.max(EnergyCounterVectorBinary) == 1:
        flag_E = 1
    else:
        flag_E = 0
    if np.max(IacCounterVectorBinary) == 1:
        flag_I = 1
    else:
        flag_I = 0
    if np.max(VdcCounterVectorBinary) == 1:
        flag_V = 1
    else:
        flag_V = 0

    return flag_E, flag_I, flag_V

def vdc_startSearchPoint(EnergyCounterVectorBinary,VdcCounterVectorBinary,IacCounterVectorBinary):
    EnergyCounterVectorBinaryARR = np.array(EnergyCounterVectorBinary)
    VdcCounterVectorBinaryARR = np.array(VdcCounterVectorBinary)
    IacCounterVectorBinaryARR = np.array(IacCounterVectorBinary)
    MultiplyAll = EnergyCounterVectorBinaryARR * VdcCounterVectorBinaryARR * IacCounterVectorBinaryARR
    MultiplyVDC_IAC =  VdcCounterVectorBinaryARR * IacCounterVectorBinaryARR
    #MultiplyVDC_Energy
    NewLabelVDC = np.zeros(len(VdcCounterVectorBinary))
    DiffEdgeVec = []
    DiffEdgeVecIDX = []
    if np.max(MultiplyAll) == 1 :
        indS1 = np.where(MultiplyAll == 1)[0][0]
        indSearch = indS1 - 40
        IndxArc = np.where(VdcCounterVectorBinaryARR[indSearch:] == 1)[0] + indSearch
        IndxArcS = IndxArc[0]
        # NewLabelVDC[IndxArcS:] = 1
        NewLabelVDC[IndxArc] = 1
        positive_edges = np.where(np.diff(NewLabelVDC) > 0)[0]
        if len(positive_edges) > 1 :
            for i in range(0,len(positive_edges)):
                DiffEdge = np.abs(positive_edges[i] - indS1)#IndxArcS)
                DiffEdgeVec.append(DiffEdge)
                DiffEdgeVecIDX.append(i)
                # CloseEdge
            #CloseEdge = np.min(DiffEdgeVec)
            myminCloseEdge = np.min(np.array(DiffEdgeVec))
            min_positions = np.where(DiffEdgeVec == myminCloseEdge)[0][0] #  [i for i, x in enumerate(DiffEdgeVec) if x == myminCloseEdge]
            NewLabelVDC = np.zeros(len(VdcCounterVectorBinary))
            S = positive_edges[min_positions] + 1
            # S = positive_edges[-1] + 1
            NewLabelVDC[S:] = 1
    elif np.max(MultiplyVDC_IAC) == 1 and np.max(EnergyCounterVectorBinaryARR) == 1 :
        indS1 = np.where(MultiplyVDC_IAC == 1)[0][0]
        indSearch = indS1 - 40
        IndxArc = np.where(VdcCounterVectorBinaryARR[indSearch:] == 1)[0] + indSearch
        # IndxArcS = IndxArc[0]
        # NewLabelVDC[IndxArcS:] = 1
        NewLabelVDC[IndxArc] = 1
        positive_edges = np.where(np.diff(NewLabelVDC) > 0)[0]
        if len(positive_edges) > 1 :
            NewLabelVDC = np.zeros(len(VdcCounterVectorBinary))
            S = positive_edges[-1] + 1
            NewLabelVDC[S:] = 1

    return NewLabelVDC

def Raises_Counter(VectorBinary):
    # Calculate the positive and negative edges
    positive_edges = np.where(np.diff(VectorBinary) > 0)[0]
    # negative_edges = np.where(np.diff(VdcCounterVectorBinary) < 0)[0]
    count = len(positive_edges) if VectorBinary[0] == 0 else len(positive_edges) + 1
    return count

def graceTimer(VdcCounterVectorBinary, VdcGraceTimerSec,IacCounterVectorBinary,IacGraceTimerSec, Fs_G,StringFile):
    try:
        VdcRaiseIDX = np.where(VdcCounterVectorBinary == 1)[0]
    except:
        VdcRaiseIDX = [] #np.nan
        print("VDC - No raise - Miss detection in rec" + StringFile)
        # MissDetect_vec.append(1)

    try:
        IacRaiseIDX = np.where(IacCounterVectorBinary == 1)[0]
    except:
        IacRaiseIDX = [] #np.nan
        print("Iac - No raise - Miss detection in rec" + StringFile)
        # MissDetect_vec.append(1)

    VdcGraceTimerSamples = int(VdcGraceTimerSec * Fs_G)
    VdcGraceTimerVec = np.zeros(len(VdcCounterVectorBinary))

    if len(VdcRaiseIDX) >= 1:
        for n in range(0, len(VdcRaiseIDX)):
            VdcGraceTimerStart = VdcRaiseIDX[n]  # np.where(np.array(VdcRaiseIDX) == 1)[0][k]
            VdcGraceTimerStop = np.min([VdcGraceTimerStart + VdcGraceTimerSamples, len(VdcCounterVectorBinary)])
            VdcGraceTimerVec[VdcGraceTimerStart:VdcGraceTimerStop] = 1

    IacGraceTimerSamples = int(IacGraceTimerSec * Fs_G)
    IacGraceTimerVec = np.zeros(len(IacCounterVectorBinary))

    if len(IacRaiseIDX) >= 1:
        for n in range(0, len(IacRaiseIDX)):
            IacGraceTimerStart = IacRaiseIDX[n]  # np.where(np.array(VdcRaiseIDX) == 1)[0][k]
            IacGraceTimerStop = np.min([IacGraceTimerStart + IacGraceTimerSamples, len(IacCounterVectorBinary)])
            IacGraceTimerVec[IacGraceTimerStart:IacGraceTimerStop] = 1

    return VdcGraceTimerVec , IacGraceTimerVec

def zcr_calcExtend(Iac_L1C,k_extendC):
    # ZCR Iac L1 - I.e., zero_crossings will contain the indices of elements before which a zero crossing occurs. If you want the elements after, just add 1 to that array.
    zero_crossings = numpy.where(numpy.diff(numpy.sign(Iac_L1C)))[0]  # .astype(int)
    zero_crossingsIdx = zero_crossings.astype(int)
    if Iac_L1C[zero_crossingsIdx[0] + 2] > 0:
        zero_crossingsIdx22 = np.array(zero_crossingsIdx[::2]).astype(int)  # [random_values.astype(int)]#zero_crossingsIdx[::2]
    else:
        zero_crossingsIdxP = zero_crossingsIdx[1:]
        zero_crossingsIdx22 = np.array(zero_crossingsIdxP[::2]).astype(int)

    DiffZCR1 = np.array(np.diff(zero_crossingsIdx22)>10, dtype=bool) #and np.diff(zero_crossingsIdx22)<400#zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    DiffZCR2 = np.array(np.diff(zero_crossingsIdx22)<400, dtype=bool) #and np.diff(zero_crossingsIdx22)<400#zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    DiffZCR = np.logical_and(DiffZCR1, DiffZCR2)
    # DiffZCR = np.diff(zero_crossingsIdx22)>10 #zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    zero_crossingsIdx2_1 = zero_crossingsIdx22[1:]
    zero_crossingsIdx2_lst = zero_crossingsIdx2_1[DiffZCR].tolist()
    zero_crossingsIdx2_lst.insert(0,zero_crossingsIdx22[0])
    zero_crossingsIdx2_lstTemp = zero_crossingsIdx2_lst.copy()
    for i in range(k_extendC):
        # zero_crossingsIdx2_lstTemp.insert(-1, zero_crossingsIdx2_lstTemp[-1]+333)
        zero_crossingsIdx2_lstTemp.extend([zero_crossingsIdx2_lstTemp[-1]+333])
        #SynthesisCountinues = zero_crossingsIdx2_lst[-1]+333
    zero_crossingsIdx2Temp = np.array(zero_crossingsIdx2_lstTemp)
    zero_crossingsIdx2 = np.array(zero_crossingsIdx2_lst)

    # DiffZCR = np.diff(zero_crossingsIdx22)>300 #zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    # zero_crossingsIdx2_1 = zero_crossingsIdx22[1:]
    # zero_crossingsIdx2_lst = zero_crossingsIdx2_1[DiffZCR].tolist()
    # zero_crossingsIdx2_lst.insert(0,zero_crossingsIdx22[0])
    # zero_crossingsIdx2 = np.array(zero_crossingsIdx2_lst)
    return zero_crossingsIdx2 , zero_crossingsIdx2Temp

def zcr_calcByPlaceInBitExtend(PlaceInBit,k_extendC):
    # ZCR Iac L1 - I.e., zero_crossings will contain the indices of elements before which a zero crossing occurs. If you want the elements after, just add 1 to that array.
    zero_crossings = numpy.where(np.array(PlaceInBit) == 0)[0]  # .astype(int)
    zero_crossingsIdx22 = zero_crossings.astype(int)
    # DiffZCR = np.diff(zero_crossingsIdx22)>300 #zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    DiffZCR1 = np.array(np.diff(zero_crossingsIdx22)>10, dtype=bool) #and np.diff(zero_crossingsIdx22)<400#zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    DiffZCR2 = np.array(np.diff(zero_crossingsIdx22)<400, dtype=bool) #and np.diff(zero_crossingsIdx22)<400#zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    DiffZCR = np.logical_and(DiffZCR1, DiffZCR2)
    # DiffZCR = np.diff(zero_crossingsIdx22)>10 #zero_crossingsIdx2[np.diff(zero_crossingsIdx2)>200]
    zero_crossingsIdx2_1 = zero_crossingsIdx22[1:]
    zero_crossingsIdx2_lst = zero_crossingsIdx2_1[DiffZCR].tolist()
    zero_crossingsIdx2_lst.insert(0,zero_crossingsIdx22[0])
    zero_crossingsIdx2_lstTemp = zero_crossingsIdx2_lst.copy()
    for i in range(k_extendC):
        # zero_crossingsIdx2_lstTemp.insert(-1, zero_crossingsIdx2_lstTemp[-1]+333)
        zero_crossingsIdx2_lstTemp.extend([zero_crossingsIdx2_lstTemp[-1]+333])
        #SynthesisCountinues = zero_crossingsIdx2_lst[-1]+333
    zero_crossingsIdx2Temp = np.array(zero_crossingsIdx2_lstTemp)
    zero_crossingsIdx2 = np.array(zero_crossingsIdx2_lst)

    return zero_crossingsIdx2 , zero_crossingsIdx2Temp


# folder = r"M:\Users\LinoyL\DATA_ARCS\DataOptimizeVDC\ReadSPI1_without_OP"
# folder = r"M:\Users\LinoyL\DATA_ARCS\DataOptimizeVDC\ReadSPI1"
folder = r"M:\Users\LinoyL\DATA_ARCS\DataOptimizeVDC\ReadSPI2\S1"




Scenario = 'S1_ByGraceTimer' #'ReadSPI1_without_OP_ByGraceTimer'
NumRec = 1500
listFolder = [x[0] for x in os.walk(folder)]
Fs_G = 50
T_G = 1 / Fs_G
Fs_SPI = 16667

T_SPI = 1 / Fs_SPI
THVdc = 0.02 #0.045
W_Size = 20# 7
F_Size = 15 # 28
CutOff_fs = 1.45 #2.5

Counter_Energy50 = 12
Counter_Iac50 = 12
Counter_Vdc50 = 3

a_filter_50 = 0.2119
TH_Energy50 = 11
k_extendC = 15

StringFile_Vec = []
FA_vec = []
MissDetect_vec = []
VdcGraceTimerSec = 1.4 #sec =  (70 Samples)
IacGraceTimerSec = 0.74 # sec = (37 Samples)

plots_per_pane = 8
fig = initialize_fig(row=8, col=1, plots_per_pane=plots_per_pane, shared_xaxes=True,
                     subplot_titles=['Rx Raw ~16.6 kHz','Energy 50 Hz','Vdc Slow 50 Hz','Iac RMS 50 Hz','Binary Detection Counter Energy 50 Hz','Binary Detection Counter Vdc Slow 50 Hz','Binary Detection Counter Iac RMS 50 Hz','Grace Timer'])

file_arr_list=list(range(1, NumRec))
for i in range(0,len(listFolder)):
    for k, filenameSPI in enumerate([f for f in os.listdir(listFolder[i]) if f.endswith('.txt') and 'spi' in f]):
        filename1SPI = f"{listFolder[i]}/{filenameSPI}"
        NameSPI = filenameSPI[:-4]
        splitFile = NameSPI.split()
        StringFileSPI = ' '.join(map(str, splitFile[:2]))
        StringFile = StringFileSPI
        # SPI1
        # Vdc_fast1, Vdc_slow50, Iac_L1, RX_RAW, lastenergy, Vac_L1 = SPI_Reading.read_file(filename1SPI)
        # SPI2
        Iac_L1, Iac_L2, Iac_L3, RX_RAW, Vdc_fast1, Vac_L1 = SPI_Reading.read_file(filename1SPI)

        zero_crossingsIdx22NoExtend, zero_crossingsIdx22 = zcr_calcExtend(Iac_L1, k_extendC)

        N_rx = len(RX_RAW)
        if zero_crossingsIdx22[-1] >= N_rx:
            zero_crossingsIdx2 = zero_crossingsIdx22[zero_crossingsIdx22 <= N_rx]
        else:
            zero_crossingsIdx2 = zero_crossingsIdx22


        indicesStop = zero_crossingsIdx2[-1]
        Vdc_fast = Vdc_fast1[:indicesStop]
        RX_RAWC = RX_RAW[:indicesStop]
        Iac_L1C = Iac_L1[:indicesStop]
        Iac_L2C = Iac_L2[:indicesStop]
        Iac_L3C = Iac_L3[:indicesStop]


        L = len(RX_RAWC)
        time_RX_RAWC = np.linspace(0, L * T_SPI, L)

        N = len(Vdc_fast)
        time_spi_fast = np.linspace(0, N * T_SPI, N)

        # Energy ZCR new 50Hz
        log_energy50, timeEnergy50 = stage1_energy_calc50_ZCR_hammingFixed(RX_RAWC, time_RX_RAWC, Fs_SPI,zero_crossingsIdx2, fIf=6000, alpha=a_filter_50, TH=TH_Energy50,
                                                    windowSize=20, filter_size=15, samTH=Counter_Energy50)
        # Vdc ZCR new 50Hz
        VdcVecFileredNew, timeNew, VecThreshV_NN, timeWindVec_NN, VdcCounterVector, timeCounterVecVdc,Vdc_Slow_New,time_Vdc_Slow_New = newVdcProcessCounterZCR(
            Vdc_fast,time_RX_RAWC,zero_crossingsIdx2, CutOff_fs, Fs_SPI, 1, windowSize=W_Size, filter_size=F_Size, samTH=Counter_Vdc50, TH_Vdc=THVdc)

        # Iac ZCR RMS 50Hz
        RmsIac_L1C, timeRmsIac_L1C = IacProcessZCR(Iac_L1C, time_RX_RAWC,zero_crossingsIdx2)
        RmsIac_L2C, timeRmsIac_L2C = IacProcessZCR(Iac_L2C, time_RX_RAWC,zero_crossingsIdx2)
        RmsIac_L3C, timeRmsIac_L3C = IacProcessZCR(Iac_L3C, time_RX_RAWC,zero_crossingsIdx2)
        RmsIacC = np.array(RmsIac_L1C) + np.array(RmsIac_L2C) + np.array(RmsIac_L3C)
        time_RmsIacC = timeRmsIac_L1C

        # Detection E
        EnergyCounterVector50, timeCounterVec50 = Counter_Energy_calcCorrect(log_energy50, timeEnergy50, buffer_size=50, windowSize=20, filter_size=15, TH_Energy=TH_Energy50)
        # Detection I
        IacCounterVector,timeCounterVecIac = Counter_Iac_calcCorrect(RmsIacC, time_RmsIacC, buffer_size=50, windowSize=20, filter_size=15, TH_Iac=0.2)
        # Binary Detection
        EnergyCounterVectorBinary = np.where(np.array(EnergyCounterVector50) >= Counter_Energy50, 1, 0)
        IacCounterVectorBinary = np.where(np.array(IacCounterVector) >= Counter_Iac50, 1, 0)
        VdcCounterVectorBinary = np.where(np.array(VdcCounterVector) >= Counter_Vdc50, 1, 0)
        # check any detection and output is a flag
        flag_E, flag_I, flag_V = E_V_I_FlagDetection(EnergyCounterVectorBinary, IacCounterVectorBinary, VdcCounterVectorBinary)

        NewClassVdc = vdc_startSearchPoint(EnergyCounterVectorBinary,VdcCounterVectorBinary,IacCounterVectorBinary)
        VdcGraceTimerVec, IacGraceTimerVec = graceTimer(VdcCounterVectorBinary, VdcGraceTimerSec,IacCounterVectorBinary, IacGraceTimerSec, Fs_G, StringFile)
        EnergyDetectedInGraceTimerVec = VdcGraceTimerVec + IacGraceTimerVec + EnergyCounterVectorBinary
        NewVec = np.zeros(len(NewClassVdc))
        NewClassRaiseIDX = np.nan
        if np.max(EnergyDetectedInGraceTimerVec) == 3:
            try:
                NewClassRaiseIDX = np.where(NewClassVdc == 1)[0][0]
                NewVec[NewClassRaiseIDX:] = 1
                Stage1PassBinary = np.where(np.array(EnergyDetectedInGraceTimerVec) == 3, 1, 0)
                Stage1PassBinaryIdx = np.where(np.array(EnergyDetectedInGraceTimerVec) == 3)[0]
                Stage1PassBinaryIdxMax = np.max(Stage1PassBinaryIdx)
                if Stage1PassBinaryIdxMax < NewClassRaiseIDX:
                    Stage1PassBinaryRaises_Counter = Raises_Counter(Stage1PassBinary)
                    Stage1PassCounterFA = np.max([Stage1PassBinaryRaises_Counter - 1, 0])
                    FA_vec.append(Stage1PassCounterFA)
                    MissDetect_vec.append(1)
                else:
                    FA_vec.append(0)
                    MissDetect_vec.append(0)
            except:
                NewVec = np.zeros(len(NewClassVdc))
        else:
            MissDetect_vec.append(1)
            FA_vec.append(0)

        StringFile_Vec.append(StringFile)


        fig.add_trace(go.Scattergl(x=time_RX_RAWC,
                                   y=RX_RAWC,
                                   name='RX_RAW' + 'Rec' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=1, col=1, secondary_y=False)

        fig.add_trace(go.Scattergl(x=timeEnergy50,
                                   y=log_energy50,
                                   name='Energy 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=2, col=1, secondary_y=False)
        fig.add_trace(go.Scattergl(x=time_Vdc_Slow_New,
                                   y=Vdc_Slow_New,
                                   name='Vdc Slow 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=3, col=1, secondary_y=False)
        fig.add_trace(go.Scattergl(x=time_RmsIacC,
                                   y=RmsIacC,
                                   name='Iac RMS 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=4, col=1, secondary_y=False)

        fig.add_trace(go.Scattergl(x=timeCounterVec50,
                                   y=EnergyCounterVectorBinary,
                                   name='Binary Detection Counter Energy 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=5, col=1, secondary_y=False)

        fig.add_trace(go.Scattergl(x=timeCounterVecVdc,
                                   y=VdcCounterVectorBinary,
                                   name='Binary Detection Counter Vdc Slow 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=6, col=1, secondary_y=False) # NewClassVdc


        fig.add_trace(go.Scattergl(x=timeCounterVecIac,
                                   y=IacCounterVectorBinary,
                                   name='Binary Detection Counter Iac RMS 50 Hz ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=7, col=1, secondary_y=False)
        fig.add_trace(go.Scattergl(x=timeCounterVecIac,
                                   y=EnergyDetectedInGraceTimerVec,
                                   name='Grace Timer ' + 'Rec ' + StringFile, mode="lines",
                                   visible=False,
                                   showlegend=True),
                      row=8, col=1, secondary_y=False)








for i in range(plots_per_pane):
    fig.data[i].visible = True
steps = []
for i in range(0, int(len(fig.data) / plots_per_pane)):
    Temp = file_arr_list[i]

    step = dict(
        method="update",
        args=[{"visible": [False] * len(fig.data)},
              {"title": "Slider  switched to step "}],
        label=str(Temp)  # layout attribute
    )
    j = i * plots_per_pane
    for k in range(plots_per_pane):
        step["args"][0]["visible"][j + k] = True
    steps.append(step)
sliders = [dict(
    active=10,
    currentvalue={"prefix": "REC: "},
    pad={"t": 50},
    steps=steps
)]
fig.update_layout(
    sliders=sliders)

config = {'scrollZoom': True, 'responsive': False, 'editable': True, 'modeBarButtonsToAdd': ['drawline',
                                                                                             'drawopenpath',
                                                                                             'drawclosedpath',
                                                                                             'drawcircle',
                                                                                             'drawrect',
                                                                                             'eraseshape'
                                                                                             ]}


fig.write_html(r"C:\Users\linoy.l\PycharmProjects\TimelineFeature\Timeline_CSV_Results\New_VDC_PARAM\0.02TH_Statistic\New_VDC_PARAM_Statistics_{}.HTML".format(Scenario), auto_open=True, config=config)


FA_SUM_Res = np.sum(FA_vec) # FP
MissDetect_SUM_Res = np.sum(MissDetect_vec) # FN
TP_Res = len(MissDetect_vec) - np.sum(MissDetect_vec) # TP
TN_Res = 0
Precision_Res = TP_Res/(TP_Res+FA_SUM_Res)
Recall_Res = TP_Res/(TP_Res+MissDetect_SUM_Res)
F1_Score_Res = 2*(Precision_Res*Recall_Res)/(Precision_Res+Recall_Res)
Accuracy_Res = (TP_Res+TN_Res)/(TP_Res+TN_Res+MissDetect_SUM_Res+FA_SUM_Res)

data = {'NameFile': StringFile_Vec,
        'FA ': FA_vec,
        'MissDetect ': MissDetect_vec
        }
df0 = pd.DataFrame(data)

writer = pd.ExcelWriter(r"C:\Users\linoy.l\PycharmProjects\TimelineFeature\Timeline_CSV_Results\New_VDC_PARAM\0.02TH_Statistic\FA_MissDetects_{}_VdcGraceTimerSec_{}_IacGraceTimerSec_{}.xlsx".format(Scenario,VdcGraceTimerSec,IacGraceTimerSec), engine="xlsxwriter")
df0.to_excel(writer, startrow=0)
writer._save()

wb_col = openpyxl.load_workbook(filename=r"C:\Users\linoy.l\PycharmProjects\TimelineFeature\Timeline_CSV_Results\New_VDC_PARAM\0.02TH_Statistic\FA_MissDetects_{}_VdcGraceTimerSec_{}_IacGraceTimerSec_{}.xlsx".format(Scenario,VdcGraceTimerSec,IacGraceTimerSec)) #worksheet# load_workbook('test.xlsx')
sheet = wb_col.active

sheet.cell(row=2, column = 7).value = "Precision .{}".format(Precision_Res)
sheet.cell(row=3, column= 7).value = "Recall .{}".format(Recall_Res)
sheet.cell(row=4, column= 7).value = "F1_Score .{}".format(F1_Score_Res)
sheet.cell(row=5, column= 7).value = "Accuracy .{}".format(Accuracy_Res)

wb_col.save(r"C:\Users\linoy.l\PycharmProjects\TimelineFeature\Timeline_CSV_Results\New_VDC_PARAM\0.02TH_Statistic\FA_MissDetects_{}_VdcGraceTimerSec_{}_IacGraceTimerSec_{}.xlsx".format(Scenario,VdcGraceTimerSec,IacGraceTimerSec))
