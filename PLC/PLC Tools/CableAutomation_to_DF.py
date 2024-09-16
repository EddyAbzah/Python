import os
import plotly
import openpyxl
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
if os.getlogin() == "eddy.a":
    from my_pyplot import omit_plot as _O, plot as _P, print_chrome as _PC, clear as _PP, print_lines as _PL, send_mail as _SM
    import Plot_Graphs_with_Sliders as _G
    import my_tools


drop_measurements = [False, -10, 100]

def split_excel_scenarios(path, sheet_name='Measurement', scenario_name='Setup of OPT', filter_cells=('power', 'arc'), stop_cells=('inv_telem', 'opt_ka')):
    """
    Gets a CableAutomation Excel file, and returns a Dictionary of all the Scenarios.
    Args:
        path: String - Full file path with extension.
        sheet_name: String - name of the main sheet; default = 'Measurement'.
        scenario_name: String - The Excel will be split to scenarios using this string; default = 'Setup of OPT'.
        filter_cells: List of strings - Parsing of each Scenario skips when encountering one of these strings; default = ['power', 'arc'].
        stop_cells: List of strings - Parsing of each Scenario stops when encountering one of these strings; default = ['inv_telem', 'opt_ka'].
    Returns:
        Dictionary of scenarios only (without samples).
    """
    print(f'split_excel_scenarios ({path = })')
    scenarios = []
    ws = openpyxl.load_workbook(path, read_only=True)[sheet_name]
    for row in ws.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value is not None and scenario_name in str(cell.value):
                if len(scenarios) > 0:
                    scenarios[-1].update({"Stop index": cell.row - 1})
                cell_string = ws.cell(row=cell.row + 1, column=1).value + ' = ' + cell.value.split(' is ')[-1]
                print(f'Scenario found at index = ({cell.row + 1}): Scenario = {cell_string}')
                # scenarios.append({"Scenario", cell_string}) is a must, the others are only for better visibility
                scenarios.append(dict([("Scenario", cell_string), ('Inverter to CB', '')] + [('String ' + s, '') for s in '123456']))
                for row_index in range(cell.row + 2, cell.row + 15):
                    cell_string = ws.cell(row=row_index, column=1).value
                    if any(s in cell_string for s in filter_cells):
                        continue
                    elif any(s in cell_string for s in stop_cells):
                        scenarios[-1].update({"Start index": row_index})
                        break
                    elif 'string' in cell_string:
                        scenarios[-1].update(dict([('String ' + cs, cell_string.split(' length ')[-1]) for cs in cell_string.split('string')[1].split(' to CB length')[0].split('+')]))
                    elif 'length' in cell_string:
                        scenarios[-1].update({cell_string.split(' length ')[0].replace('inv', 'Inv'): cell_string.split(' length ')[-1]})
                    else:   # KA + Telem freqs
                        scenarios[-1].update({cell_string.split(': ')[0]: cell_string.split(': ')[-1][:2] + 'kHz'})
    if "Stop index" not in scenarios[-1]:
        scenarios[-1].update({"Stop index": ws.max_row - 1})
    print(f'Total Scenarios = {len(scenarios)}')
    print()
    return scenarios


def excel_to_df(path, scenarios, sheet_name='Measurement', skip_table=('inv_telem_rssi', 'opt_ka_rssi')):
    """
    Gets a CableAutomation Excel file with a Scenarios Dictionary, and returns a DataFrame of the whole test.
    Args:
        path: String - Full file path with extension.
        scenarios: Dictionary - the Dictionary output from split_excel_scenarios().
        sheet_name: String - name of the main sheet; default = 'Measurement'.
        skip_table: List of strings - Skip these table titles; default = ['inv_telem_rssi', 'opt_ka_rssi'].
    Returns:
        DataFrame.
    """
    print(f'excel_to_df (len of scenarios = {len(scenarios)})')
    all_dfs = pd.DataFrame()
    for scenario in scenarios:
        print(f'{scenario = }')
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, skiprows=scenario["Start index"] - 1, nrows=scenario["Stop index"] - scenario["Start index"])
        df = df.dropna(axis=1)
        split_indexes = [i for i, v in enumerate(df.iloc[:, 1]) if isinstance(v, str)]
        for current_index, (start_index, stop_index) in enumerate(zip(split_indexes, split_indexes[1:] + [len(df)])):
            table_title = df.iloc[start_index, 0]
            print(f'Current table is = {table_title}')
            if any(s == table_title for s in skip_table):
                print('This table is skipped')
                continue
            sdf = pd.concat([pd.DataFrame({'Measurement': table_title}, index=df[start_index + 1:stop_index].index).join(df[start_index + 1:stop_index])], axis=0)
            sdf.columns = ['Measurement', 'Sample'] + ['OPT_' + ''.join([c for c in s if c.isdigit()]) for s in df.iloc[start_index, 1:]]
            if drop_measurements[0]:
                sdf = sdf.map(lambda x: x if isinstance(x, str) or (drop_measurements[1] < x < drop_measurements[2]) else np.nan)
            sdf.reset_index(drop=True, inplace=True)
            all_dfs = pd.concat([all_dfs, pd.DataFrame({k: v for k, v in scenario.items() if 'index' not in k}, index=sdf.index).join(sdf)], axis=0)
        print()
    all_dfs.reset_index(drop=True, inplace=True)
    return all_dfs


if __name__ == "__main__":
    path_folder = r'M:\Users\HW Infrastructure\PLC team\INVs\Jupiter48\Jupiter48 BU - New layout + DC conducted - EddyA 2.2024\Cable Automation\Cable Automation 05'
    path_file_in = 'Jup48 New DC Filter.xlsx'
    path_file_out = 'Jup48 New DC Filter'
    scenarios = split_excel_scenarios(f'{path_folder}\\{path_file_in}')
    pd.DataFrame(scenarios).to_csv(f'{path_folder}\\{path_file_out} - Scenarios.csv', index=False)
    df = excel_to_df(path_folder + '\\' + path_file_in, scenarios)
    df.to_csv(f'{path_folder}\\{path_file_out}.csv', index=False)
    print()
